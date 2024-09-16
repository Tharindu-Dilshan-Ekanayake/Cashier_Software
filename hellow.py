import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Function to add items to the bill and save to Excel
def add_item():
    item_name = item_name_entry.get()
    item_price = item_price_entry.get()

    if item_name and item_price:
        try:
            price = float(item_price)
            bill_list.insert(tk.END, f"{item_name} - ${price:.2f}")
            total_price[0] += price
            total_label.config(text=f"Total: ${total_price[0]:.2f}")
            save_to_excel(item_name, price)
            item_name_entry.delete(0, tk.END)
            item_price_entry.delete(0, tk.END)
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid price.")
    else:
        messagebox.showerror("Missing Input", "Please fill in both item name and price.")

# Function to save data to an Excel file
def save_to_excel(item_name, price):
    file_name = "cashier_transactions.xlsx"

    if os.path.exists(file_name):
        # Load existing workbook
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        # Create new workbook
        workbook = Workbook()
        sheet = workbook.active
        # Add headers to the Excel file
        sheet.append(["Item Name", "Price"])

    # Append the transaction data
    sheet.append([item_name, price])

    # Save the workbook
    workbook.save(file_name)

# Function to clear the bill
def clear_bill():
    bill_list.delete(0, tk.END)
    total_price[0] = 0
    total_label.config(text="Total: $0.00")

# Create the main window
root = tk.Tk()
root.title("Cashier Software")

# Initialize total price as a list to make it mutable
total_price = [0]

# Create a label for item name
tk.Label(root, text="Item Name:", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=10)
item_name_entry = tk.Entry(root, font=("Arial", 12))
item_name_entry.grid(row=0, column=1, padx=10, pady=10)

# Create a label for item price
tk.Label(root, text="Item Price:", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10)
item_price_entry = tk.Entry(root, font=("Arial", 12))
item_price_entry.grid(row=1, column=1, padx=10, pady=10)

# Create a button to add the item
add_button = tk.Button(root, text="Add Item", font=("Arial", 12), command=add_item)
add_button.grid(row=2, column=0, columnspan=2, pady=10)

# Create a listbox to display the items and their prices
bill_list = tk.Listbox(root, font=("Arial", 12), width=40, height=10)
bill_list.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

# Create a label to display the total price
total_label = tk.Label(root, text="Total: $0.00", font=("Arial", 14))
total_label.grid(row=4, column=0, columnspan=2, pady=10)

# Create a button to clear the bill
clear_button = tk.Button(root, text="Clear Bill", font=("Arial", 12), command=clear_bill)
clear_button.grid(row=5, column=0, columnspan=2, pady=10)

# Start the Tkinter event loop
root.mainloop()
